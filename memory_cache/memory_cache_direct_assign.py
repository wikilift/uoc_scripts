import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
def direct_assign(memory_accesses,block_size=4,num_lines=4,
                  initial_state=True):
    e=0
    def calculate_line_and_tag(address, block_size, num_lines):
        block_number = address // block_size
        tag = block_number // num_lines
        line = block_number % num_lines
        return block_number, tag, line

    cache = {i: (i, 0) for i in range(num_lines)}

    cache_evolution = pd.DataFrame(columns=['Estado Inicial'] + [f'Acceso {addr}' for addr in memory_accesses])
    if initial_state:
        for line in range(num_lines):
            block_number, tag = cache[line]
            start_word = block_number * block_size
            cache_evolution.at[line, 'Estado Inicial'] = f"{block_number}:{tag} ({start_word}, {start_word + 1}, {start_word + 2}, {start_word + 3})"

    for address in memory_accesses:
        block_number, tag, line = calculate_line_and_tag(address, block_size, num_lines)
        access_label = f'Acceso {address}'
        
        if cache[line][0] == block_number:

            cache_evolution.at[line, access_label] = 'Acierto'
            e+=1
        else:
            
            cache_evolution.at[line, access_label] = f"{block_number}:{tag} ({block_number*block_size}-{block_number*block_size + 3})"
            cache[line] = (block_number, tag)

   

    hit_ratio = e / len(memory_accesses)
    time_succes = 4  
    time_fault = 20  
    tm = time_succes * hit_ratio + time_fault * (1 - hit_ratio)
    print(f"La tasa de aciertos (Hit Ratio) es: {hit_ratio}")
    print(f"El tiempo de acceso a memoria medio es: {tm}ns")

        # Incrementamos el contador LRU para todas las líneas antes de verificar si el bloque está en caché
    cache_evolution.fillna('', inplace=True)
    cache_evolution.to_excel("output/cache_evolution.xlsx")  
    excel_file = "output/cache_evolution.xlsx"
    cache_evolution.fillna('', inplace=True)
    cache_evolution.to_excel(excel_file, index=False)
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    bold_font = Font(bold=True)
    sheet.cell(row=7, column=1, value=f"La tasa de aciertos (Hit Ratio) es: {hit_ratio:.4f}").font = bold_font
    sheet.cell(row=8, column=1, value=f"El tiempo de acceso a memoria medio es: {tm:.2f} ns").font = bold_font

    workbook.save(excel_file)
    
    print("Done!")



memory_accesses=[1, 2, 13, 14, 26, 27, 28, 29, 36, 37, 38, 39, 40, 3, 10, 11, 12, 13, 14, 15, 30, 8, 12]
block_size=4
num_lines=4

direct_assign(initial_state=True,
              block_size=block_size,
              num_lines=num_lines,
              memory_accesses=memory_accesses)